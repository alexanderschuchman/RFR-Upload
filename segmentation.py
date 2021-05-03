import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import win32com.client as win32
from pathlib import Path
import os
import re
import sys
import shutil
# import win32com

def sortFile():
    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
    except AttributeError:
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        excel = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        wb = excel.Workbooks("C:/Users/bp_bhagyashree phadn/Downloads/RFR-Upload/samples/RFRSample.xlsx")
    except Exception as e:
        try:
            wb = excel.Workbooks.Open("C:/Users/bp_bhagyashree phadn/Downloads/RFR-Upload/samples/RFRSample.xlsx")
        except Exception as e:
            print(e)
    try:
        ws = wb.Worksheets('Sheet1')
        
        LastRow = ws.UsedRange.Rows.Count
        
        ws.Range('A2:R'+str(LastRow)).Sort(Key1=ws.Range('R1'), Order1=1, Orientation=1)
        
        wb.SaveAs(Filename="C:\\Users\\bp_bhagyashree phadn\\Downloads\\RFR-Upload\\samples\\sorted.xlsx")
        # wb.Save()
        excel.Application.Quit()
    except Exception as e:
        print(e)

def generateGroups():
    wb1 = load_workbook("samples/sorted.xlsx")
    ws1 = wb1.worksheets[0]
    groups = {}
    i = 2
    val = ws1.cell(row = 2, column = 18).value
    # print(ws1.max_row)
    while i <= ws1.max_row:
        while ws1.cell(row = i, column = 18).value == val:
            if val not in groups.keys():
                groups[val] = [ws1.cell(row = i, column = 3).value]
                # print(groups)
                # print(i)
            elif ws1.cell(row = i, column = 3).value not in groups[val]:
                groups[val].append(ws1.cell(row = i, column = 3).value)
                # print(groups)
                # print(i)
            i += 1
        if ws1.cell(row = i, column = 18).value != val:
            val = ws1.cell(row = i, column = 18).value
            i += 1
    # print(groups)
    # groups['Discontinued/ Obselete'] += groups['Discontinued/ Obselete']
    # groups['Discontinued/ Obselete'] += groups['Discontinued/ Obselete']
    # groups['Discontinued/ Obselete'] += groups['Discontinued/ Obselete']
    for k in groups.keys():
        if len(groups[k])>200:
            groups[k] = [groups[k][:200], groups[k][200:]]
            while len(groups[k][-1])>200:
                excess = groups[k][-1][200:]
                groups[k][-1] = groups[k][-1][:200]
                groups[k].append(excess)
    print(groups)
    return groups

def saveExcel(groups):
    book = Workbook()
    sheet = book.worksheets[0]
    col_count = 1
    for k in groups.keys():
        sheet.cell(row = 1, column = col_count).value = k
        count = 2
        if isinstance(groups[k][0], list):
            for i in range(0, len(groups[k])):
                for j in groups[k][i]:
                    sheet.cell(row = count, column = col_count).value = j
                    count += 1
                if i!=len(groups[k])-1:
                    col_count += 1
                    sheet.cell(row = 1, column = col_count).value = k
                    count = 2
        else:
            for i in groups[k]:
                sheet.cell(row = count, column = col_count).value = i
                count += 1
        col_count += 1
    book.save("samples/output.xlsx")


sortFile()
groups = generateGroups()
saveExcel(groups)